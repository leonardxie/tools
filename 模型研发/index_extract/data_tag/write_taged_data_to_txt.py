#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@Project ：tools 
@File    ：write_taged_data_to_txt.py.py
@IDE     ：PyCharm 
@Author  ：解书贵
@Date    ：2024/5/11 14:18 
'''

# 写入xlsx文件
from openpyxl import Workbook, load_workbook
import json


def main():
    # 读取xlsx文件
    work_book = load_workbook('data.xlsx')

    # 选择工作表
    sheet = work_book.active

    text_data_list = []
    labeled_tag_list = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        cleaned_row = [cell for cell in row if cell is not None]
        if not cleaned_row:
            break
        if row_idx % 2 == 1:  # 判断是奇数行还是偶数行
            text_data_list.append(cleaned_row)
        else:
            labeled_tag_list.append(cleaned_row)
    work_book.close()

    # 正样本数据 写入json文件中
    with open('../corpus/train.txt', "w", encoding='utf-8') as fw:
        for i in range(len(text_data_list)):
            item = {'text': text_data_list[i], 'label': labeled_tag_list[i]}
            json_data = json.dumps(item, ensure_ascii=False)
            fw.write(json_data + "\n")

    print('success')


if __name__ == '__main__':
    main()